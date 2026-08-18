import re
import unittest

import openpyxl
from openpyxl.cell.cell import MergedCell

from tenforty.mappings.f1040 import F1040
from tests.helpers import SPREADSHEETS_DIR


class TestF1040Inputs2025(unittest.TestCase):
    def test_has_2025_inputs(self):
        inputs = F1040.get_inputs(2025)
        self.assertIsInstance(inputs, dict)
        self.assertGreater(len(inputs), 0)

    def test_w2_wage_fields(self):
        inputs = F1040.get_inputs(2025)
        self.assertEqual(inputs["w2_wages_1"], "C3")
        self.assertEqual(inputs["w2_fed_withheld_1"], "C4")
        self.assertEqual(inputs["w2_ss_wages_1"], "C5")
        self.assertEqual(inputs["w2_ss_withheld_1"], "C6")
        self.assertEqual(inputs["w2_medicare_wages_1"], "C7")
        self.assertEqual(inputs["w2_medicare_withheld_1"], "C8")

    def test_filing_status_fields(self):
        inputs = F1040.get_inputs(2025)
        self.assertEqual(inputs["filing_status_single"], "File_Single")
        self.assertEqual(inputs["filing_status_married_jointly"], "File_Marr_Joint")
        self.assertEqual(inputs["filing_status_married_separately"], "File_Marr_Sep")
        self.assertEqual(inputs["filing_status_head_of_household"], "File_Head")

    def test_birthdate_fields(self):
        inputs = F1040.get_inputs(2025)
        self.assertEqual(inputs["birthdate_month"], "YourBirthMonth")
        self.assertEqual(inputs["birthdate_day"], "YourBirthDay")
        self.assertEqual(inputs["birthdate_year"], "YourBirthYear")

    def test_1099_int_fields(self):
        inputs = F1040.get_inputs(2025)
        self.assertIn("interest_1", inputs)

    def test_1098_mortgage_interest(self):
        inputs = F1040.get_inputs(2025)
        self.assertIn("mortgage_interest", inputs)

    def test_schedule_e_rental_fields(self):
        inputs = F1040.get_inputs(2025)
        self.assertIn("sche_rents_a", inputs)
        self.assertIn("sche_property_type_a", inputs)


class TestF1040Outputs2025(unittest.TestCase):
    def test_has_2025_outputs(self):
        outputs = F1040.get_outputs(2025)
        self.assertIsInstance(outputs, dict)
        self.assertGreater(len(outputs), 0)

    def test_core_output_fields(self):
        outputs = F1040.get_outputs(2025)
        self.assertEqual(outputs["agi"], "Adj_Gross_Inc")
        self.assertEqual(outputs["taxable_income"], "Taxable_Inc")
        # `total_tax` is IRS 1040 line 16 (`Tax_SubTotal`) on every path. It
        # formerly pointed at `Tax`, which is line 18 = line 16 + Schedule 2
        # Part I. See tenforty/mappings/f1040.py OUTPUTS for the rationale and
        # TestF1040TaxBandOutputsEveryYear for the all-years pin.
        self.assertEqual(outputs["total_tax"], "Tax_SubTotal")
        self.assertEqual(outputs["federal_withheld"], "W2_FedTaxWH")
        self.assertEqual(outputs["overpaid"], "Overpaid")

    def test_schedule_e_output(self):
        outputs = F1040.get_outputs(2025)
        self.assertEqual(outputs["sche_line26"], "SchE1_Line26")

    def test_total_income_output(self):
        outputs = F1040.get_outputs(2025)
        self.assertEqual(outputs["total_income"], "Total_Income")

    def test_total_payments_output(self):
        outputs = F1040.get_outputs(2025)
        self.assertEqual(outputs["total_payments"], "Tot_Payments")

    def test_total_deductions_output(self):
        outputs = F1040.get_outputs(2025)
        self.assertEqual(outputs["total_deductions"], "TotalDeductions")

    def test_schedule_a_total_output(self):
        outputs = F1040.get_outputs(2025)
        self.assertEqual(outputs["schedule_a_total"], "Tot_Item_Deduct")

    def test_standard_deduction_uses_filing_status_aware_range(self):
        outputs = F1040.get_outputs(2025)
        self.assertEqual(
            outputs["standard_deduction"], "Standard",
            "standard_deduction must map to Standard (1040!BI70, the filing-status-aware "
            "dollar amount), not SD_Single (single-only) or StdDeduct (boolean flag).",
        )


class TestF1040TaxBandOutputsEveryYear(unittest.TestCase):
    """The 1040 tax band (lines 16/17/18) pinned across ALL FIVE supported
    years, not just 2025. These assertions are deliberately NOT in
    `TestF1040Outputs2025`: 2023/2022/2021 are produced by `F1040.inherit`
    off the 2024 block rather than written out literally, so a per-year loop
    is what proves the inheritance chain actually carries these keys."""

    YEARS = (2021, 2022, 2023, 2024, 2025)

    def test_total_tax_is_line_16_on_every_year(self):
        """`total_tax` means IRS 1040 line 16 on every path. The workbook
        previously mapped it to `Tax` (line 18 = line 16 + Schedule 2 Part I),
        which printed an overstated line 16 and double-counted excess APTC
        through f1040x line 6. Pinned per-year: 2023/2022/2021 inherit these
        OUTPUTS, so a regression in the inheritance chain surfaces here."""
        for year in self.YEARS:
            with self.subTest(year=year):
                self.assertEqual(
                    F1040.get_outputs(year)["total_tax"], "Tax_SubTotal")

    def test_total_tax_line16_is_absent_from_every_year(self):
        """`total_tax_line16` must not exist on any year. It was a DUPLICATE
        of `total_tax` — both pointed at the `Tax_SubTotal` named range and
        named the same quantity, IRS 1040 line 16.

        Its absence is PINNED rather than assumed because a duplicate key is
        the mechanism by which the semantic fork silently returns: while
        `total_tax` meant line 16 on the native spine and line 18 on the
        workbook, this key existed solely so the parity battery could compare
        like with like. Two names for one quantity means someone can repoint
        one and not the other, and tenforty is back to one key with two
        meanings — with every path still self-consistent and no test noticing,
        which is exactly how the original fork survived.

        Checked on all five years' RESOLVED outputs, not just the two explicit
        OUTPUTS blocks: 2023 inherits 2024, 2022 inherits 2023 and 2021
        inherits 2022 via `F1040.inherit` (a plain dict-merge that can add but
        never remove), so re-adding the key to either explicit block would
        propagate down the chain. The per-year loop is what proves that.
        """
        for year in self.YEARS:
            with self.subTest(year=year):
                self.assertNotIn(
                    "total_tax_line16", F1040.get_outputs(year),
                    f"{year} carries `total_tax_line16`, a retired duplicate "
                    f"of `total_tax`. Both name IRS 1040 line 16; use "
                    f"`total_tax`. For line 17 use `schedule2_tax`, for line "
                    f"18 `tax_plus_schedule2`, for line 24 "
                    f"`tax_liability_line24`.",
                )

    def test_schedule2_and_line18_keys_have_workbook_producers(self):
        """Both keys are mapped to PDF boxes in all five year blocks of
        pdf_1040.py but had NO producer on either path, so lines 17 and 18
        printed blank on every emitted 1040. This wires the workbook side."""
        for year in self.YEARS:
            with self.subTest(year=year):
                outputs = F1040.get_outputs(year)
                self.assertEqual(outputs["schedule2_tax"], "Schedule2_Tax")
                self.assertEqual(outputs["tax_plus_schedule2"], "Tax")

    def test_deduction_diagnostic_is_harvested_by_name_every_year(self):
        """The refusal guard is only as good as its harvest.

        `forms/f1040.py::workbook_refusal` passes silently when this key is
        ABSENT -- it must, since a missing key is indistinguishable from a
        caller that does not harvest it. So a year that lost the mapping
        would silently return to reading a refused blank as a number, and no
        refusal test would notice. This is the assertion that notices.

        Mapped to the NAME `Deduction`, never to an address: the cell moves
        every year (2021 '1040'!AJ51, 2022 AJ58, 2023 AK64, 2024 AK67, 2025
        AU91) while the name is stable in all five workbooks. Note the name
        is SINGULAR -- `Deductions` (plural) is a DIFFERENT named range, the
        line-12 dollar amount, which is the value the diagnostic zeroes.
        """
        for year in self.YEARS:
            with self.subTest(year=year):
                self.assertEqual(
                    F1040.get_outputs(year)["deduction_diagnostic"],
                    "Deduction",
                )

    def test_deduction_named_range_exists_in_every_workbook(self):
        """The mapping VALUE being right is not the same as the NAME existing.

        This is the assertion that keeps the refusal guard from quietly
        becoming dead code while its own tests certify it — the exact failure
        mode this unit exists to remove, turned on the guard itself.

        `oracle/engine.py::_read_outputs` resolves an OUTPUT by looking its
        value up in `wb.defined_names`, falling back to SHEET_MAP, and then
        SILENTLY assigning None (`engine.py:192`). So if a future year's
        workbook ships without a `Deduction` name, the key is still PRESENT in
        the harvest, carrying None; `forms/f1040.py::workbook_refusal` reads
        that as "no diagnostic" and passes — by design, since an absent
        diagnostic must not refuse — and refused blanks get read as data
        again, with every non-oracle test still green and the sibling
        `test_deduction_diagnostic_is_harvested_by_name_every_year` above
        still passing, because the MAPPING would be untouched.

        Asserting resolution against the real workbooks is the only thing that
        catches that. Same shape as `TestF1040Form8962Mapping::
        test_ptc_output_named_ranges` and `TestF1040EstimatedTaxPayments::
        test_named_range_resolves_in_every_workbook` below.
        """
        for year in self.YEARS:
            with self.subTest(year=year):
                workbook_path = (
                    SPREADSHEETS_DIR / "federal" / str(year) / "1040.xlsx"
                )
                wb = openpyxl.load_workbook(workbook_path, read_only=False)
                try:
                    self.assertIn(
                        "Deduction", wb.defined_names,
                        f"{year} is missing the `Deduction` named range. The "
                        f"harvest would fall through to None "
                        f"(oracle/engine.py:192), silently disarming the "
                        f"workbook-refusal guard in forms/f1040.py. Do NOT "
                        f"repoint this at a cell address — the address drifts "
                        f"every year. Find what the vendor renamed it to.",
                    )
                    # And it must be a SINGLE-AREA name: exactly one
                    # destination entry. That is ALL this establishes — it does
                    # NOT prove the name reaches a real cell. One entry can be
                    # a multi-cell RANGE, and openpyxl yields an entry just as
                    # happily for a sheet that does not exist. The existence
                    # assertion above is the load-bearing one; this only rules
                    # out the name growing extra areas.
                    destinations = list(
                        wb.defined_names["Deduction"].destinations)
                    self.assertEqual(
                        len(destinations), 1,
                        f"{year} `Deduction` must have exactly one "
                        f"destination entry",
                    )
                finally:
                    wb.close()


def _sole_destination(wb, name: str) -> tuple[str, str]:
    """Return the single (sheet, cell) a defined name resolves to."""
    destinations = list(wb.defined_names[name].destinations)
    if len(destinations) != 1:
        raise AssertionError(
            f"`{name}` must resolve to exactly one destination; "
            f"got {destinations!r}")
    sheet_name, cell_ref = destinations[0]
    return sheet_name, cell_ref.replace("$", "")


class TestF1040TotalTaxLiabilityLine24(unittest.TestCase):
    """1040 LINE 24 (`Tot_Tax`), harvested for Form 4868 line 4.

    Form 4868 line 4 is "Estimate of Total Tax Liability" and its instruction
    (pdfs/federal/2025/f4868.pdf) names 1040 line 24 explicitly. Line 24 is
    NOT `total_tax` (line 16) and NOT `tax_plus_schedule2` (line 18), so it
    needs its own OUTPUT — see forms/f4868.py for why the workbook path
    HARVESTS it while the native path composes.

    2023/2022/2021 inherit these OUTPUTS off 2024 rather than declaring them,
    so the per-year loop is what proves the inheritance chain carries the key.
    """

    YEARS = (2021, 2022, 2023, 2024, 2025)

    # Line 24's cell in each shipped workbook, for the formula-shape pin
    # below. Recorded so a vendor re-layout is visible as a test failure
    # naming the year, not as a silent semantic change.
    LINE_24_CELL = {
        2021: ("1040", "AC65"),
        2022: ("1040", "AC74"),
        2023: ("1040", "AD80"),
        2024: ("1040", "AD83"),
        2025: ("1040", "AL104"),
    }

    def test_mapped_to_the_name_on_every_year(self):
        for year in self.YEARS:
            with self.subTest(year=year):
                self.assertEqual(
                    F1040.get_outputs(year)["tax_liability_line24"], "Tot_Tax")

    def test_named_range_exists_in_every_workbook(self):
        """The mapping VALUE being right is not the same as the NAME existing.

        `oracle/engine.py::_read_outputs` resolves an OUTPUT by looking its
        value up in `wb.defined_names` and SILENTLY assigns None
        (`engine.py:192`) when the name does not resolve. A `Tot_Tax` that
        stopped resolving would therefore harvest as None, f4868 would fall
        through to its native composition on the WORKBOOK path, and line 4
        would quietly lose NIIT and AMT with every mapping test still green.
        Same shape as `TestF1040Form8962Mapping::test_ptc_output_named_ranges`.
        """
        for year in self.YEARS:
            with self.subTest(year=year):
                workbook_path = (
                    SPREADSHEETS_DIR / "federal" / str(year) / "1040.xlsx"
                )
                wb = openpyxl.load_workbook(workbook_path, read_only=False)
                try:
                    self.assertIn(
                        "Tot_Tax", wb.defined_names,
                        f"{year} is missing the `Tot_Tax` named range, so "
                        f"Form 4868 line 4 would silently fall back to the "
                        f"incomplete native composition. Do NOT repoint this "
                        f"at a cell address — line 24's address moves in "
                        f"every one of the five workbooks.",
                    )
                    self.assertEqual(
                        self.LINE_24_CELL[year],
                        _sole_destination(wb, "Tot_Tax"),
                    )
                finally:
                    wb.close()

    def test_line_24_is_line_22_plus_line_23_with_the_floor_on_line_22(self):
        """Pin the vendor's own arithmetic, including WHERE the floor sits.

        Every year:  line 24 = SUM(line 22, line 23)
                     line 22 = MAX(0, SUM(Tax, -<line 21 credits>))
                     line 23 = TotalOtherTaxes   (Schedule 2 Part II)

        The zero floor is on (line 18 - nonrefundable credits) and nothing
        else — Schedule 2 Part II is added AFTER it. forms/f4868.py's native
        composition mirrors that order; this is the evidence it mirrors.
        """
        for year in self.YEARS:
            with self.subTest(year=year):
                workbook_path = (
                    SPREADSHEETS_DIR / "federal" / str(year) / "1040.xlsx"
                )
                wb = openpyxl.load_workbook(workbook_path, data_only=False)
                try:
                    sheet_name, cell_ref = _sole_destination(wb, "Tot_Tax")
                    sheet = wb[sheet_name]
                    line_24 = sheet[cell_ref].value
                    match = re.fullmatch(
                        r"=SUM\(([A-Z]+\d+),\s*([A-Z]+\d+)\)", str(line_24))
                    self.assertIsNotNone(
                        match,
                        f"{year} line 24 ({cell_ref}) is no longer a plain "
                        f"two-operand SUM: {line_24!r}",
                    )
                    line_22_ref, line_23_ref = match.group(1), match.group(2)

                    line_22 = str(sheet[line_22_ref].value)
                    self.assertIn(
                        "MAX(0,SUM(Tax,-", line_22.replace(" ", ""),
                        f"{year} line 22 ({line_22_ref}) no longer floors "
                        f"(line 18 - credits) at zero: {line_22!r}",
                    )
                    # The floor's operand is line 18 MINUS credits and stops
                    # there: Schedule 2 Part II must NOT be inside it.
                    self.assertNotIn(
                        "TotalOtherTaxes", line_22,
                        f"{year} line 22 ({line_22_ref}) now floors Schedule "
                        f"2 Part II too, which would swallow other taxes a "
                        f"filer owes: {line_22!r}",
                    )
                    line_23 = str(sheet[line_23_ref].value)
                    self.assertIn(
                        "TotalOtherTaxes", line_23,
                        f"{year} line 23 ({line_23_ref}) is no longer "
                        f"Schedule 2 Part II: {line_23!r}",
                    )
                finally:
                    wb.close()


class TestF1040MappingValidity(unittest.TestCase):
    """Pre-flight checks: every direct cell ref in SHEET_MAP must point at a
    writable cell in the actual workbook. Catches merged-cell mapping bugs
    before they surface as cryptic 'MergedCell attribute is read-only' errors
    during e2e runs."""

    def test_no_input_maps_to_merged_cell(self):
        for year, sheet_map in F1040.SHEET_MAP.items():
            workbook_path = SPREADSHEETS_DIR / "federal" / str(year) / "1040.xlsx"
            if not workbook_path.exists():
                continue
            wb = openpyxl.load_workbook(workbook_path, read_only=False)
            inputs = F1040.get_inputs(year)
            outputs = F1040.get_outputs(year)
            for key, sheet_name in sheet_map.items():
                # SHEET_MAP serves both input writes and output reads (the
                # engine falls back to SHEET_MAP in _write_inputs and
                # _read_outputs alike). Merged-cell concern only applies
                # to writes — reading a merged cell is safe.
                if key not in inputs:
                    self.assertIn(
                        key, outputs,
                        f"{year} SHEET_MAP key '{key}' is in neither INPUTS nor OUTPUTS",
                    )
                    continue
                cell_ref = inputs[key]
                cell = wb[sheet_name][cell_ref]
                self.assertNotIsInstance(
                    cell, MergedCell,
                    f"{year} input '{key}' maps to {sheet_name}!{cell_ref}, "
                    f"which is a merged cell. Map to the top-left of the merge range instead.",
                )


class TestF1040InputTypes(unittest.TestCase):
    def test_all_input_values_are_strings(self):
        for key, value in F1040.get_inputs(2025).items():
            self.assertIsInstance(value, str, f"Input '{key}' value is {type(value)}, expected str")

    def test_all_output_values_are_strings(self):
        for key, value in F1040.get_outputs(2025).items():
            self.assertIsInstance(value, str, f"Output '{key}' value is {type(value)}, expected str")


class TestF1040Form8949Mapping(unittest.TestCase):
    """Form 8949 per-lot row mapping: 4 boxes (A/B/D/E) × 11 lots × 7 fields.

    Box C/F (no 1099-B) is out of scope since Form1099B implies a received
    1099-B by definition.
    """

    def test_each_box_has_eleven_lot_rows(self):
        inputs = F1040.get_inputs(2025)
        for box in ("a", "b", "d", "e"):
            for idx in range(1, 12):
                self.assertIn(
                    f"f8949_box_{box}_lot_{idx}_description", inputs,
                    f"missing Box {box.upper()} lot {idx} description cell",
                )

    def test_box_a_lot_1_cells_are_part_i_row_41(self):
        inputs = F1040.get_inputs(2025)
        self.assertEqual(inputs["f8949_box_a_lot_1_description"],     "AJ41")
        self.assertEqual(inputs["f8949_box_a_lot_1_date_acquired"],   "AK41")
        self.assertEqual(inputs["f8949_box_a_lot_1_date_sold"],       "AL41")
        self.assertEqual(inputs["f8949_box_a_lot_1_proceeds"],        "AM41")
        self.assertEqual(inputs["f8949_box_a_lot_1_basis"],           "AN41")
        self.assertEqual(inputs["f8949_box_a_lot_1_adjustment_code"], "AO41")
        self.assertEqual(inputs["f8949_box_a_lot_1_adjustment_amount"], "AP41")

    def test_box_d_lot_1_cells_are_part_ii_row_91(self):
        inputs = F1040.get_inputs(2025)
        self.assertEqual(inputs["f8949_box_d_lot_1_description"], "AJ91")
        self.assertEqual(inputs["f8949_box_d_lot_1_proceeds"],    "AM91")

    def test_box_a_and_d_use_sheet_8949A(self):
        sheet_map = F1040.SHEET_MAP[2025]
        self.assertEqual(sheet_map["f8949_box_a_lot_1_description"], "8949A")
        self.assertEqual(sheet_map["f8949_box_d_lot_1_description"], "8949A")

    def test_box_b_and_e_use_sheet_8949B(self):
        sheet_map = F1040.SHEET_MAP[2025]
        self.assertEqual(sheet_map["f8949_box_b_lot_1_description"], "8949B")
        self.assertEqual(sheet_map["f8949_box_e_lot_1_description"], "8949B")

    def test_box_totals_map_to_named_ranges(self):
        """Per-box totals come from the workbook's 4-letter named ranges:
        {sheet}{S,L}T{D,E,G,H} where D/E/G/H columns are proceeds/basis/
        adjustment/gain respectively on the subsection total row."""
        outputs = F1040.get_outputs(2025)
        self.assertEqual(outputs["f8949_box_a_total_proceeds"],   "F8949ASTD")
        self.assertEqual(outputs["f8949_box_a_total_gain"],       "F8949ASTH")
        self.assertEqual(outputs["f8949_box_b_total_proceeds"],   "F8949BSTD")
        self.assertEqual(outputs["f8949_box_d_total_proceeds"],   "F8949ALTD")
        self.assertEqual(outputs["f8949_box_e_total_gain"],       "F8949BLTH")


class TestF1040Form8962Mapping(unittest.TestCase):
    """Form 8962 (Premium Tax Credit) monthly-grid input mapping.

    Recon (docs/plans/f8962-probe-tables.md, 'Workbook 8962-tab layout'):
    monthly-grid INPUT columns are fixed across all 5 years — (a) enrollment
    premium = G, (b) SLCSP = L, (f) APTC = AF — on the '8962' sheet. Only the
    ROW range drifts by year. All 5 years verified blank/writable; a
    write-then-readback (openpyxl, no save/recalc) proves it per test run
    rather than trusting the recon doc alone.
    """

    YEAR_ROWS: dict[int, range] = {
        2021: range(52, 64),
        2022: range(48, 60),
        2023: range(48, 60),
        2024: range(48, 60),
        2025: range(50, 62),
    }
    MONTH_COLS = (("premium", "G"), ("slcsp", "L"), ("aptc", "AF"))

    def test_monthly_grid_cells_are_blank_and_writable(self):
        for year, rows in self.YEAR_ROWS.items():
            with self.subTest(year=year):
                inputs = F1040.get_inputs(year)
                sheet_map = F1040.SHEET_MAP[year]
                workbook_path = SPREADSHEETS_DIR / "federal" / str(year) / "1040.xlsx"
                wb = openpyxl.load_workbook(workbook_path, read_only=False)
                sheet = wb["8962"]
                for n, row in zip(range(1, 13), rows):
                    for field, col in self.MONTH_COLS:
                        key = f"f8962_month_{n}_{field}"
                        self.assertIn(key, inputs, f"{year} missing input key {key}")
                        self.assertEqual(inputs[key], f"{col}{row}")
                        self.assertIn(key, sheet_map, f"{year} missing SHEET_MAP entry for {key}")
                        self.assertEqual(sheet_map[key], "8962")
                        cell = sheet[inputs[key]]
                        self.assertNotIsInstance(
                            cell, MergedCell,
                            f"{year} input '{key}' maps to 8962!{inputs[key]}, "
                            f"which is a merged cell.",
                        )
                        self.assertIsNone(
                            cell.value,
                            f"{year} input '{key}' cell 8962!{inputs[key]} is not blank "
                            f"(found {cell.value!r}) — not writable as expected.",
                        )
                        marker = f"__TEST_{key}__"
                        cell.value = marker
                        self.assertEqual(sheet[inputs[key]].value, marker)
                wb.close()

    def test_2021_ui_checkbox(self):
        inputs = F1040.get_inputs(2021)
        sheet_map = F1040.SHEET_MAP[2021]
        self.assertEqual(inputs["f8962_ui_checkbox"], "AI14")
        self.assertEqual(sheet_map["f8962_ui_checkbox"], "8962")
        workbook_path = SPREADSHEETS_DIR / "federal" / "2021" / "1040.xlsx"
        wb = openpyxl.load_workbook(workbook_path, read_only=False)
        cell = wb["8962"]["AI14"]
        self.assertNotIsInstance(cell, MergedCell)
        self.assertIsNone(cell.value, f"2021 f8962_ui_checkbox cell AI14 is not blank (found {cell.value!r}).")
        cell.value = "X"
        self.assertEqual(wb["8962"]["AI14"].value, "X")
        wb.close()

    def test_ui_checkbox_only_wired_for_2021(self):
        for year in (2022, 2023, 2024, 2025):
            with self.subTest(year=year):
                self.assertNotIn("f8962_ui_checkbox", F1040.get_inputs(year))

    def test_ptc_output_named_ranges(self):
        for year in self.YEAR_ROWS:
            with self.subTest(year=year):
                outputs = F1040.get_outputs(year)
                self.assertEqual(outputs["f8962_net_ptc"], "PTC_Net")
                self.assertEqual(outputs["f8962_repayment"], "PTC_Excess")
                workbook_path = SPREADSHEETS_DIR / "federal" / str(year) / "1040.xlsx"
                wb = openpyxl.load_workbook(workbook_path, read_only=False)
                self.assertIn("PTC_Net", wb.defined_names, f"{year} missing PTC_Net named range")
                self.assertIn("PTC_Excess", wb.defined_names, f"{year} missing PTC_Excess named range")
                wb.close()


class TestF1040EstimatedTaxPayments(unittest.TestCase):
    """Estimated tax payments (Form 1040 line 26) input: a named range,
    EstimatedTaxPayments, present in every year's workbook pointing at that
    year's own line-26 cell. Named ranges resolve per-workbook automatically,
    so no SHEET_MAP entry is needed."""

    def test_named_range_mapping_every_year(self):
        for year in (2021, 2022, 2023, 2024, 2025):
            with self.subTest(year=year):
                inputs = F1040.get_inputs(year)
                self.assertEqual(inputs["estimated_tax_payments"], "EstimatedTaxPayments")

    def test_named_range_resolves_in_every_workbook(self):
        for year in (2021, 2022, 2023, 2024, 2025):
            with self.subTest(year=year):
                workbook_path = SPREADSHEETS_DIR / "federal" / str(year) / "1040.xlsx"
                wb = openpyxl.load_workbook(workbook_path, read_only=False)
                self.assertIn(
                    "EstimatedTaxPayments", wb.defined_names,
                    f"{year} missing EstimatedTaxPayments named range",
                )
                wb.close()


class TestF1040CharitableNonitemizer2021(unittest.TestCase):
    """2021 line 12b (CARES/CAA above-the-line cash-charitable deduction for
    non-itemizers): a named range, Charitable, present ONLY in the 2021
    workbook (the provision expired — 2022-2025 workbooks lack the range),
    so this input mapping is 2021-scoped and deliberately NOT inherited into
    later years."""

    def test_2021_mapping(self):
        inputs = F1040.get_inputs(2021)
        self.assertEqual(inputs["charitable_nonitemizer"], "Charitable")

    def test_named_range_resolves_in_2021_workbook(self):
        workbook_path = SPREADSHEETS_DIR / "federal" / "2021" / "1040.xlsx"
        wb = openpyxl.load_workbook(workbook_path, read_only=False)
        self.assertIn(
            "Charitable", wb.defined_names,
            "2021 missing Charitable named range",
        )
        wb.close()

    def test_not_mapped_in_later_years(self):
        for year in (2022, 2023, 2024, 2025):
            with self.subTest(year=year):
                inputs = F1040.get_inputs(year)
                self.assertNotIn("charitable_nonitemizer", inputs)
