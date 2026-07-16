import os
import shutil
import subprocess
import tempfile
import threading
import unittest.mock
from pathlib import Path

import openpyxl


# Belt-and-suspenders on top of per-invocation UserInstallation (see #23).
# Serializes every soffice invocation so concurrent callers can't race on
# any soffice-internal shared resource (sockets, registry caches) that
# profile isolation doesn't cover.
_SOFFICE_LOCK = threading.Lock()


def _resolve_named_range(defn: object) -> tuple[str, str]:
    """Parse a named range definition into (sheet_name, cell_address)."""
    dest = defn.value
    sheet_name, cell_addr = dest.split("!")
    sheet_name = sheet_name.strip("'")
    cell_addr = cell_addr.replace("$", "")
    return sheet_name, cell_addr


def _assert_oracle_sanctioned() -> None:
    """Fail loudly if soffice/LibreOffice is reached from a non-oracle-marked test.

    Inert in production: PYTEST_CURRENT_TEST is unset outside pytest, so this never
    fires and soffice runs normally. Under pytest, an oracle-marked test has its
    sanction set by the conftest hookwrapper (see tests/conftest.py); an unmarked
    test that routes here has no sanction -> RuntimeError. Guards the LAUNCH, so it
    catches ANY route to soffice, orchestrator or otherwise.
    """
    current = os.environ.get("PYTEST_CURRENT_TEST")
    if not current:
        return  # production: PYTEST_CURRENT_TEST unset -> soffice runs normally
    if os.environ.get("TENFORTY_ORACLE_SANCTIONED") == "1":
        return  # oracle-marked test: sanctioned by the conftest hookwrapper
    if isinstance(subprocess.run, unittest.mock.NonCallableMock):
        # subprocess.run has been replaced by a mock (e.g. tests/test_oracle_engine.py,
        # which A3's AST entry-point guard already allowlists for the same reason): no
        # REAL soffice can launch, so there is nothing to guard. NOTE: a Mock whose
        # side_effect delegates to the real subprocess.run could evade this exemption;
        # that is OUT of the threat model -- the tripwire targets ACCIDENTAL leaks, not
        # deliberate circumvention (the whole-branch review and A3's AST guard cover
        # that). Do NOT tighten this into something that re-breaks the engine's
        # mock-based unit tests.
        return
    raise RuntimeError(
        f"soffice/LibreOffice was reached from a non-oracle-marked test "
        f"({current}); decorate that test @needs_libreoffice (oracle-tier). "
        f"See the tiering-leak remediation."
    )


class SpreadsheetEngine:
    """Writes inputs into a spreadsheet, recalculates via LibreOffice, reads outputs."""

    def compute(
        self,
        spreadsheet_path: Path,
        mapping: type,
        year: int,
        inputs: dict[str, object],
        work_dir: Path | None = None,
    ) -> dict[str, object]:
        input_map = mapping.get_inputs(year)
        output_map = mapping.get_outputs(year)
        sheet_map = getattr(mapping, "SHEET_MAP", {}).get(year, {})

        work_dir = work_dir or Path("/tmp/tenforty_work")
        work_dir.mkdir(parents=True, exist_ok=True)

        working_copy = work_dir / spreadsheet_path.name
        shutil.copy2(spreadsheet_path, working_copy)

        self._write_inputs(working_copy, input_map, sheet_map, inputs)
        recalculated = self._recalculate(working_copy, work_dir)
        return self._read_outputs(recalculated, output_map, sheet_map)

    def _write_inputs(
        self,
        workbook_path: Path,
        input_map: dict[str, str],
        sheet_map: dict[str, str],
        inputs: dict[str, object],
    ) -> None:
        wb = openpyxl.load_workbook(workbook_path)
        named_ranges = {n.name: n for n in wb.defined_names.values()}

        for input_key, value in inputs.items():
            if input_key not in input_map:
                continue

            cell_ref = input_map[input_key]

            if cell_ref in named_ranges:
                defn = named_ranges[cell_ref]
                sheet_name, cell_addr = _resolve_named_range(defn)
                wb[sheet_name][cell_addr] = value
            elif input_key in sheet_map:
                sheet_name = sheet_map[input_key]
                wb[sheet_name][cell_ref] = value
            else:
                raise ValueError(
                    f"Input '{input_key}' maps to '{cell_ref}' but has no named range "
                    f"and no sheet in SHEET_MAP"
                )

        wb.save(workbook_path)

    def _recalculate(self, workbook_path: Path, work_dir: Path) -> Path:
        _assert_oracle_sanctioned()
        output_dir = work_dir / "recalculated"
        output_dir.mkdir(exist_ok=True)
        expected_output = output_dir / workbook_path.name

        # Per-invocation UserInstallation sidesteps the profile lock at
        # ~/.config/libreoffice/4/.~lock.registrymodifications.xcu# so
        # concurrent soffice frontends can't silently exit 0 without
        # producing output.
        with _SOFFICE_LOCK, tempfile.TemporaryDirectory(prefix="soffice_profile_") as profile_dir:
            try:
                result = subprocess.run(
                    [
                        "soffice",
                        f"-env:UserInstallation=file://{profile_dir}",
                        "--headless", "--calc",
                        "--convert-to", "xlsx",
                        "--outdir", str(output_dir),
                        str(workbook_path),
                    ],
                    capture_output=True,
                    text=True,
                    timeout=60,
                )
            # TimeoutExpired is NOT CalledProcessError — would leak past the
            # returncode check. Re-raise as RuntimeError so downstream callers
            # see a uniform error surface.
            except subprocess.TimeoutExpired as e:
                raise RuntimeError(
                    f"soffice timeout after {e.timeout}s for "
                    f"{expected_output}; stdout={e.stdout!r} stderr={e.stderr!r}"
                ) from e

        if result.returncode != 0:
            raise RuntimeError(
                f"soffice recalculation failed (exit={result.returncode}): "
                f"stderr={result.stderr!r} stdout={result.stdout!r}"
            )
        # soffice can exit 0 without creating output when the profile lock at
        # ~/.config/libreoffice/4/.~lock.registrymodifications.xcu# is held by
        # a concurrent invocation. The per-invocation UserInstallation above
        # sidesteps that lock; this check is residual defense.
        if not expected_output.exists():
            raise RuntimeError(
                f"soffice exited 0 but did not create {expected_output}. "
                f"stdout={result.stdout!r} stderr={result.stderr!r}"
            )
        # Zero-byte or truncated output passes .exists() but fails downstream
        # in openpyxl.load_workbook with a confusing BadZipFile error. Catch
        # the empty case here; openpyxl handles truncation-but-nonempty.
        if expected_output.stat().st_size == 0:
            raise RuntimeError(
                f"soffice exited 0 and created {expected_output} but it is empty. "
                f"stdout={result.stdout!r} stderr={result.stderr!r}"
            )
        return expected_output

    def _read_outputs(
        self,
        workbook_path: Path,
        output_map: dict[str, str],
        sheet_map: dict[str, str],
    ) -> dict[str, object]:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        named_ranges = {n.name: n for n in wb.defined_names.values()}
        results: dict[str, object] = {}

        for output_key, cell_ref in output_map.items():
            if cell_ref in named_ranges:
                defn = named_ranges[cell_ref]
                sheet_name, cell_addr = _resolve_named_range(defn)
                results[output_key] = wb[sheet_name][cell_addr].value
            elif output_key in sheet_map:
                sheet_name = sheet_map[output_key]
                results[output_key] = wb[sheet_name][cell_ref].value
            else:
                results[output_key] = None

        return results
